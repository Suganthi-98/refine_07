"""
Workbook Parser

Reads Excel workbook and converts sheets into domain models.
Structure: Row 1 = Title, Row 2 = Headers, Row 3+ = Data
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
import logging
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import uuid

logger = logging.getLogger(__name__)

from app.domain.models import (
    ProjectInfo,
    Resource,
    Sprint,
    WorkItem,
    Dependency,
    Blocker,
    SprintActual,
    ProjectState,
    SkillLevel,
    WorkItemType,
    Priority,
    WorkItemStatus,
    SprintStatus,
    BlockerSeverity,
    BlockerStatus,
    BlockerCategory,
    DependencyType,
)

BLOCKER_CATEGORY_MAP = {
    "vendor": BlockerCategory.VENDOR,
    "hardware": BlockerCategory.HARDWARE,
    "specification": BlockerCategory.SPECIFICATION,
    "resource": BlockerCategory.RESOURCE,
    "environment": BlockerCategory.ENVIRONMENT,
    "security": BlockerCategory.SECURITY,
    "compliance": BlockerCategory.COMPLIANCE,
    "lab issue": BlockerCategory.LAB_ISSUE,
    "hardware / procurement": BlockerCategory.HARDWARE_PROCUREMENT,
    "hardware/procurement": BlockerCategory.HARDWARE_PROCUREMENT,
    "external team dependency": BlockerCategory.EXTERNAL_TEAM_DEPENDENCY,
    "awaiting validation": BlockerCategory.AWAITING_VALIDATION,
    "awaiting validation from central team": BlockerCategory.AWAITING_VALIDATION,
    "tool issue": BlockerCategory.TOOL_ISSUE,
    "license unavailable": BlockerCategory.LICENSE_UNAVAILABLE,
    "license not available": BlockerCategory.LICENSE_UNAVAILABLE,
    "people dependency": BlockerCategory.PEOPLE_DEPENDENCY,
    "approval pending": BlockerCategory.APPROVAL_PENDING,
}


def parse_blocker_category(raw) -> BlockerCategory:
    if not raw:
        return BlockerCategory.OTHER
    return BLOCKER_CATEGORY_MAP.get(str(raw).strip().lower(), BlockerCategory.OTHER)


class WorkbookParseError(Exception):
    """Raised when workbook parsing fails."""
    pass


class WorkbookParser:
    """
    Parses Excel workbook into ProjectState.
    
    Each sheet has:
    - Row 1: Title (skip)
    - Row 2: Column headers
    - Row 3+: Data rows
    """
    
    REQUIRED_SHEETS = [
        "Project_Info", "Team", "Sprint_Plan", "Work_Items",
        "Dependencies", "Blockers"
    ]
    
    def __init__(self, file_path: str):
        """Initialize parser with Excel file path."""
        self.file_path = file_path
        self.workbook = None
        self.project_id = str(uuid.uuid4())[:8]
    
    def parse(self) -> ProjectState:
        """
        Parse workbook and return ProjectState.
        
        Returns:
            ProjectState: Parsed and validated project state
            
        Raises:
            WorkbookParseError: If parsing fails
        """
        try:
            # Load cached cell values so formula-derived workbook columns are parsed numerically.
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            
            # Verify all required sheets exist
            self._verify_sheets()
            
            # Parse each sheet
            project_info = self._parse_project_info()
            team = self._parse_team()
            # Keep parsed team on the parser instance for lookups when resolving work item owners
            self.team = team
            sprint_plan_rows = self._get_sheet_data("Sprint_Plan")
            sprints = self._parse_sprints(sprint_plan_rows)
            work_items = self._parse_work_items()
            dependencies = self._parse_dependencies()
            blockers = self._parse_blockers()
            actuals = self._build_sprint_actuals(sprints, work_items, sprint_plan_rows)
            
            # Create and return ProjectState
            return ProjectState(
                project_id=self.project_id,
                project_info=project_info,
                team=team,
                sprints=sprints,
                work_items=work_items,
                dependencies=dependencies,
                blockers=blockers,
                actuals=actuals,
            )
        except Exception as e:
            raise WorkbookParseError(f"Failed to parse workbook: {str(e)}") from e
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _verify_sheets(self) -> None:
        """Verify all required sheets exist."""
        sheet_names = self.workbook.sheetnames
        for sheet in self.REQUIRED_SHEETS:
            if sheet not in sheet_names:
                raise WorkbookParseError(f"Required sheet '{sheet}' not found")
    
    def _get_sheet_data(self, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Get data rows from a sheet as list of dicts.
        
        Assumes:
        - Row 1: Title (skip)
        - Row 2: Headers
        - Row 3+: Data
        """
        ws = self.workbook[sheet_name]
        
        # Row 2 contains headers
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(2, col).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        if not headers:
            raise WorkbookParseError(f"No headers found in sheet '{sheet_name}' at row 2")
        
        # Row 3+ contains data
        data_rows = []
        for row_idx in range(3, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col_idx, header in enumerate(headers, start=1):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value is not None:
                    has_data = True
                row_data[header] = cell_value
            
            # Only include rows with at least some data
            if has_data:
                data_rows.append(row_data)
        
        return data_rows
    
    def _parse_project_info(self) -> ProjectInfo:
        """Parse Project_Info sheet (single row)."""
        data_rows = self._get_sheet_data("Project_Info")
        
        if not data_rows:
            raise WorkbookParseError("Project_Info sheet has no data")
        
        row = data_rows[0]
        
        return ProjectInfo(
            project_name=self._get_str(row, "Project Name"),
            sponsor=self._get_str(row, "Sponsor"),
            business_unit=self._get_str(row, "Business Unit"),
            project_manager=self._get_str(row, "Project Manager"),
            start_date=self._get_datetime(row, "Start Date"),
            release_date=self._get_optional_datetime(row, "Release Date"),
            target_end_date=self._get_datetime(row, "Target End Date"),
            sprint_duration_days=self._get_int(row, "Sprint Length (Days)"),
            methodology=self._get_str(row, "Methodology"),
            customer=self._get_str(row, "Customer"),
            status=self._get_str(row, "Status"),
        )
    
    def _parse_team(self) -> List[Resource]:
        """Parse Team sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Team")
        resources = []
        
        for row in data_rows:
            resource_name = row.get("Resource Name")
            if not resource_name or str(resource_name).strip().lower().startswith("skill level"):
                continue
            resource_name = self._get_str(row, "Resource Name")
            
            # Generate resource ID from name (use first name initials + last name)
            resource_id = self._generate_resource_id(resource_name)
            
            resources.append(Resource(
                resource_id=resource_id,
                name=resource_name,
                role=self._get_str(row, "Role"),
                primary_skill=self._get_str(row, "Skill 1"),
                secondary_skill=self._get_optional_str(row, "Skill 2"),
                skill_level=self._parse_skill_level(row, "Skill 1 Level"),
                allocation_pct=self._average_pct_columns(row, "Alloc %"),
                availability_pct=self._average_pct_columns(row, "Avail %"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return resources
    
    def _parse_sprints(self, data_rows: List[Dict[str, Any]]) -> List[Sprint]:
        """Parse Sprint_Plan sheet (multiple rows)."""
        sprints = []
        sprint_number = 0
        
        for row in data_rows:
            # Skip rows that don't have Sprint Name (e.g., summary sections)
            sprint_name = row.get("Sprint Name")
            if not sprint_name or not str(sprint_name).strip().lower().startswith("sprint"):
                continue
            
            sprint_number += 1
            sprint_name = self._get_str(row, "Sprint Name")
            sprint_id = self._generate_sprint_id(sprint_name, sprint_number)
            
            start_date = self._get_datetime(row, "Start Date")
            end_date = self._get_datetime(row, "End Date")
            working_days = self._get_int(row, "Duration (Days)")
            
            sprints.append(Sprint(
                sprint_id=sprint_id,
                sprint_name=sprint_name,
                sprint_number=sprint_number,
                start_date=start_date,
                end_date=end_date,
                working_days=working_days,
                sprint_goal=self._get_str(row, "Sprint Goal"),
                status=self._parse_sprint_status(row),
                planned_velocity_hrs=self._get_float_safe(row, "Velocity (h)"),
                carryover_count=self._get_int(row, "Carry-Over Items"),
            ))
        
        return sprints
    
    def _derive_carryover_metrics(
        self,
        work_items: List[WorkItem],
        sprints: List[Sprint],
    ) -> Dict[str, Dict[str, float]]:
        """Derive historical carry-over metrics from work item planning history."""
        sprint_by_name = {
            self._normalize_sprint_name(sprint.sprint_name): sprint.sprint_id
            for sprint in sprints
        }
        metrics = {
            sprint.sprint_id: {
                "carry_out_count": 0,
                "carry_in_count": 0,
                "carry_out_hours": 0.0,
                "carry_in_hours": 0.0,
            }
            for sprint in sprints
        }

        for work_item in work_items:
            original_sprint = self._normalize_sprint_name(work_item.original_sprint)
            assigned_sprint = self._normalize_sprint_name(work_item.assigned_sprint)

            if not original_sprint or not assigned_sprint:
                continue

            if original_sprint == assigned_sprint:
                continue

            if original_sprint not in sprint_by_name or assigned_sprint not in sprint_by_name:
                continue

            origin_id = sprint_by_name[original_sprint]
            assigned_id = sprint_by_name[assigned_sprint]

            metrics[origin_id]["carry_out_count"] += 1
            metrics[origin_id]["carry_out_hours"] += work_item.current_estimate_hrs
            metrics[assigned_id]["carry_in_count"] += 1
            metrics[assigned_id]["carry_in_hours"] += work_item.current_estimate_hrs

        return metrics

    def _derive_sprint_history_metrics(
        self,
        work_items: List[WorkItem],
        sprints: List[Sprint],
    ) -> Dict[str, Dict[str, Any]]:
        """Derive additional sprint history metrics from work items."""
        carry_metrics = self._derive_carryover_metrics(work_items, sprints)
        sprint_by_name = {
            self._normalize_sprint_name(sprint.sprint_name): sprint.sprint_id
            for sprint in sprints
        }

        metrics = {
            sprint.sprint_id: {
                "planned_items": 0,
                "completed_items": 0,
                "planned_hours": 0.0,
                "actual_hours": 0.0,
                "carry_out_count": carry_metrics[sprint.sprint_id]["carry_out_count"],
                "carry_in_count": carry_metrics[sprint.sprint_id]["carry_in_count"],
                "carry_out_hours": carry_metrics[sprint.sprint_id]["carry_out_hours"],
                "carry_in_hours": carry_metrics[sprint.sprint_id]["carry_in_hours"],
            }
            for sprint in sprints
        }

        for work_item in work_items:
            original_sprint = self._normalize_sprint_name(work_item.original_sprint)
            assigned_sprint = self._normalize_sprint_name(work_item.assigned_sprint)

            if original_sprint and original_sprint in sprint_by_name:
                original_id = sprint_by_name[original_sprint]
                metrics[original_id]["planned_items"] += 1
                metrics[original_id]["planned_hours"] += work_item.estimated_effort_hrs

            if assigned_sprint and assigned_sprint in sprint_by_name:
                assigned_id = sprint_by_name[assigned_sprint]
                metrics[assigned_id]["actual_hours"] += work_item.actual_effort_hrs
                if work_item.status in {WorkItemStatus.DONE, WorkItemStatus.COMPLETED}:
                    metrics[assigned_id]["completed_items"] += 1

        for sprint_metrics in metrics.values():
            sprint_metrics["completion_rate"] = (
                sprint_metrics["completed_items"] / sprint_metrics["planned_items"]
                if sprint_metrics["planned_items"]
                else 0.0
            )
            sprint_metrics["velocity_ratio"] = (
                sprint_metrics["actual_hours"] / sprint_metrics["planned_hours"]
                if sprint_metrics["planned_hours"]
                else 0.0
            )
            sprint_metrics["variance_hours"] = (
                sprint_metrics["actual_hours"] - sprint_metrics["planned_hours"]
            )

        return metrics

    def _build_sprint_actuals(
        self,
        sprints: List[Sprint],
        work_items: List[WorkItem],
        sprint_plan_rows: List[Dict[str, Any]],
    ) -> List[SprintActual]:
        """Build SprintActual objects using derived sprint history and optional actuals sheet data."""
        sprint_metrics = self._derive_sprint_history_metrics(work_items, sprints)
        plan_row_map = {
            self._normalize_sprint_name(row.get("Sprint Name")): row
            for row in sprint_plan_rows
            if row.get("Sprint Name")
        }
        actual_rows = []
        if "Sprint_Actuals" in self.workbook.sheetnames:
            actual_rows = self._get_sheet_data("Sprint_Actuals")
        actual_row_map = {
            self._normalize_sprint_name(row.get("Sprint")): row
            for row in actual_rows
            if row.get("Sprint")
        }

        actuals = []
        for sprint in sprints:
            sprint_id = sprint.sprint_id
            sprint_name = sprint.sprint_name
            metrics = sprint_metrics[sprint_id]
            plan_row = plan_row_map.get(self._normalize_sprint_name(sprint_name))
            actual_row = actual_row_map.get(self._normalize_sprint_name(sprint_name))

            planned_effort_hrs = (
                self._get_float_safe(actual_row, "Planned Hours")
                if actual_row and "Planned Hours" in actual_row
                else metrics["planned_hours"]
            )
            actual_effort_hrs = (
                self._get_float_safe(actual_row, "Actual Hours")
                if actual_row and "Actual Hours" in actual_row
                else metrics["actual_hours"]
            )
            variance_hrs = (
                self._get_float_safe(actual_row, "Variance (h)")
                if actual_row and "Variance (h)" in actual_row
                else actual_effort_hrs - planned_effort_hrs
            )
            tasks_planned = (
                int(actual_row["Tasks Planned"])
                if actual_row and "Tasks Planned" in actual_row and actual_row["Tasks Planned"] is not None
                else metrics["planned_items"]
            )
            tasks_completed = (
                int(actual_row["Tasks Completed"])
                if actual_row and "Tasks Completed" in actual_row and actual_row["Tasks Completed"] is not None
                else metrics["completed_items"]
            )
            completion_rate = (
                self._get_float_safe(actual_row, "Completion Rate")
                if actual_row and "Completion Rate" in actual_row
                else (
                    tasks_completed / tasks_planned if tasks_planned else 0.0
                )
            )
            completion_rate = min(max(completion_rate, 0.0), 1.0)
            carry_out_count = metrics["carry_out_count"]
            carry_in_count = metrics["carry_in_count"]
            carry_out_hours = metrics["carry_out_hours"]
            carry_in_hours = metrics["carry_in_hours"]

            if plan_row and "Carry-Over Items" in plan_row:
                try:
                    sprint_plan_carry = int(plan_row["Carry-Over Items"] or 0)
                except (ValueError, TypeError):
                    sprint_plan_carry = 0

                if sprint_plan_carry != carry_out_count:
                    logger.warning(
                        "\nWARNING\n\nSprint %s\n\nSprint_Plan carry-over = %s\n\nDerived carry-over = %s\n\nPlease verify workbook consistency.",
                        sprint_name,
                        sprint_plan_carry,
                        carry_out_count,
                    )

            scope_change_hours = (
                self._get_float_safe(actual_row, "Scope Change Hours")
                if actual_row and "Scope Change Hours" in actual_row
                else 0.0
            )
            blocker_impact_hrs = (
                self._get_float_safe(actual_row, "Blocker Impact (h)")
                if actual_row and "Blocker Impact (h)" in actual_row
                else 0.0
            )
            notes = self._get_optional_str(actual_row, "Notes") if actual_row else None

            actuals.append(SprintActual(
                sprint_id=sprint_id,
                sprint_number=sprint.sprint_number,
                planned_effort_hrs=planned_effort_hrs,
                actual_effort_hrs=actual_effort_hrs,
                variance_hrs=variance_hrs,
                tasks_planned=tasks_planned,
                tasks_completed=tasks_completed,
                completion_rate=completion_rate,
                carryover_count=carry_out_count,
                carry_out_count=carry_out_count,
                carry_in_count=carry_in_count,
                carry_out_hours=carry_out_hours,
                carry_in_hours=carry_in_hours,
                scope_change_hours=scope_change_hours,
                blocker_impact_hrs=blocker_impact_hrs,
                notes=notes,
            ))

        return actuals

    def _parse_work_items(self) -> List[WorkItem]:
        """Parse Work_Items sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Work_Items")
        work_items = []
        
        for row in data_rows:
            # Skip rows that don't have Task ID (e.g., summary/totals rows)
            item_id = row.get("Task ID")
            if not item_id or not str(item_id).strip().upper().startswith("WI-"):
                continue
            
            item_id = self._get_str(row, "Task ID")
            
            progress_pct = self._parse_progress_pct(row)
            current_estimate = self._get_float_safe(row, "Curr Est (h)")
            original_estimate = self._get_float_safe(row, "Orig Est (h)")
            estimated_effort = original_estimate if original_estimate > 0.0 else current_estimate

            # Resolve Owner -> resource_id where possible. The workbook's "Owner" column
            # historically contains a human display name; convert that to the generated
            # `resource_id` when a matching team member exists so downstream engines
            # can rely on `WorkItem.assigned_resource` being an id.
            owner_raw = self._get_optional_str(row, "Owner")
            assigned_resource_id = None
            if owner_raw:
                # If the team has already been parsed, prefer exact id match, then name match,
                # then generated slug match (to tolerate slightly different spellings).
                try:
                    team_list = getattr(self, "team", []) or []
                except Exception:
                    team_list = []

                owner_norm = str(owner_raw).strip()
                # exact id match
                if any(getattr(r, "resource_id", None) == owner_norm for r in team_list):
                    assigned_resource_id = owner_norm
                else:
                    # name match
                    matched = next((r for r in team_list if getattr(r, "name", None) == owner_norm), None)
                    if matched:
                        assigned_resource_id = matched.resource_id
                    else:
                        # fallback: generate slug and see if that matches an id
                        slug = self._generate_resource_id(owner_norm)
                        if any(getattr(r, "resource_id", None) == slug for r in team_list):
                            assigned_resource_id = slug

            work_items.append(WorkItem(
                item_id=item_id,
                title=self._get_str(row, "Task Name"),
                work_type=self._parse_work_item_type(row),
                assigned_sprint=self._normalize_sprint_name(self._get_str(row, "Sprint")),
                original_sprint=self._normalize_sprint_name(
                    self._get_optional_str(row, "Orig. Sprint")
                ),
                assigned_resource=assigned_resource_id,
                required_skill=self._get_str(row, "Required Skill"),
                priority=self._parse_priority(row),
                estimated_effort_hrs=estimated_effort,
                current_estimate_hrs=current_estimate,
                actual_effort_hrs=self._get_float_safe(row, "Actual Hrs"),
                progress_pct=progress_pct,
                remaining_effort_hrs=self._resolve_remaining_effort(
                    row.get("Remaining Hrs"),
                    current_estimate,
                    self._parse_work_item_status(row),
                    progress_pct=progress_pct,
                ),
                status=self._parse_work_item_status(row),
                is_scope_changed=self._parse_yes_no(self._get_optional_str(row, "Scope Change")),
                scope_change_reason=self._get_optional_str(row, "Scope Reason"),
            ))
        
        return work_items
    
    def _resolve_remaining_effort(
        self,
        remaining_value: Any,
        current_estimate: float,
        status: WorkItemStatus,
        progress_pct: float = 0.0,
    ) -> float:
        """Resolve remaining effort based on status, workbook value, and progress."""
        if status in {WorkItemStatus.DONE, WorkItemStatus.COMPLETED}:
            return 0.0

        if remaining_value is not None:
            if isinstance(remaining_value, str) and remaining_value.strip() == "":
                remaining_value = None
            else:
                try:
                    return float(remaining_value)
                except (ValueError, TypeError):
                    return 0.0

        if status == WorkItemStatus.NOT_STARTED:
            return current_estimate

        if progress_pct > 0.0:
            return max(0.0, current_estimate * (1.0 - progress_pct))

        return current_estimate

    def _parse_dependencies(self) -> List[Dependency]:
        """Parse Dependencies sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Dependencies")
        dependencies = []
        
        for row in data_rows:
            successor_key = "Successor Task" if "Successor Task" in row else "Sucessor Task"
            dependencies.append(Dependency(
                dependency_id=self._get_str(row, "Dep ID"),
                predecessor_item_id=self._get_str(row, "Predecessor Task"),
                successor_item_id=self._get_str(row, successor_key),
                dependency_type=self._parse_dependency_type(row),
                is_on_critical_path=self._parse_yes_no(self._get_optional_str(row, "Critical Path")),
                lag_days=self._get_int(row, "Lag Days"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return dependencies
    
    def _parse_blockers(self) -> List[Blocker]:
        """Parse Blockers sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Blockers")
        blockers = []
        
        for row in data_rows:
            impacted_str = self._get_str(row, "Impacted Task IDs")
            impacted_ids = [x.strip() for x in impacted_str.split(",")]
            
            # Use notes as description if available, otherwise construct from other fields.
            # The Blockers sheet uses "Notes / Escalation Path" as the column header;
            # fall back to plain "Notes" for workbooks that use the shorter variant.
            notes_text = (
                self._get_optional_str(row, "Notes / Escalation Path")
                or self._get_optional_str(row, "Notes")
            )
            description = notes_text or f"Blocker {self._get_str(row, 'Blocker ID')}: {self._get_str(row, 'Related Task')}"

            blockers.append(Blocker(
                blocker_id=self._get_str(row, "Blocker ID"),
                related_item_id=self._get_str(row, "Related Task"),
                impacted_item_ids=impacted_ids,
                description=description,
                severity=self._parse_blocker_severity(row),
                status=self._parse_blocker_status(row),
                owner=self._get_optional_str(row, "Owner"),
                raised_date=self._get_datetime(row, "Raised Date"),
                target_resolution_date=self._get_optional_datetime(row, "Target Resolution"),
                actual_resolution_date=self._get_optional_datetime(row, "Actual Resolution"),
                category=parse_blocker_category(row.get("Category")),
                notes=notes_text,
            ))
        
        return blockers
    
    def _parse_sprint_actuals(self) -> List[SprintActual]:
        """Parse Sprint_Actuals sheet (multiple rows)."""
        data_rows = self._get_sheet_data("Sprint_Actuals")
        actuals = []
        
        for sprint_num, row in enumerate(data_rows, start=1):
            sprint_name = self._get_str(row, "Sprint")
            sprint_id = self._generate_sprint_id(sprint_name, sprint_num)
            
            actuals.append(SprintActual(
                sprint_id=sprint_id,
                sprint_number=sprint_num,
                planned_effort_hrs=self._get_float(row, "Planned Hours"),
                actual_effort_hrs=self._get_float(row, "Actual Hours"),
                variance_hrs=self._get_float_safe(row, "Variance (h)"),
                tasks_planned=self._get_int(row, "Tasks Planned"),
                tasks_completed=self._get_int(row, "Tasks Completed"),
                completion_rate=self._get_float_safe(row, "Completion Rate"),
                carryover_count=self._get_int(row, "Carry-Over Count"),
                scope_change_hours=self._get_float(row, "Scope Change Hours"),
                blocker_impact_hrs=self._get_float(row, "Blocker Impact (h)"),
                notes=self._get_optional_str(row, "Notes"),
            ))
        
        return actuals
    
    # ─── Helper Methods ──────────────────────────────────────────────────────
    
    def _get_str(self, row: Dict, key: str) -> str:
        """Get required string value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        return str(value).strip()
    
    def _get_optional_str(self, row: Dict, key: str) -> Optional[str]:
        """Get optional string value from row."""
        value = row.get(key)
        if value is None:
            return None
        result = str(value).strip()
        return result if result else None
    
    def _get_int(self, row: Dict, key: str) -> int:
        """Get required integer value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        try:
            return int(value)
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid integer for field {key}: {value}")
    
    def _get_float(self, row: Dict, key: str) -> float:
        """Get required float value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        try:
            return float(value)
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid float for field {key}: {value}")
    
    def _get_float_safe(self, row: Dict, key: str) -> float:
        """Get float value, return 0.0 if missing or invalid (for formulas)."""
        value = row.get(key)
        if value is None:
            return 0.0
        if isinstance(value, str) and value.startswith("="):
            # Formula string, can't evaluate
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _parse_progress_pct(self, row: Dict) -> float:
        """Parse Progress % as a decimal between 0.0 and 1.0."""
        value = row.get("Progress %")
        if value is None:
            return 0.0
        if isinstance(value, str) and value.startswith("="):
            return 0.0
        try:
            progress_value = float(value)
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid float for field Progress %: {value}")

        if progress_value < 0.0 or progress_value > 100.0:
            raise WorkbookParseError(
                f"Invalid Progress % value: {progress_value}. Expected 0-100 or 0.0-1.0."
            )

        if progress_value > 1.0:
            progress_value /= 100.0

        return min(max(progress_value, 0.0), 1.0)

    def _get_datetime(self, row: Dict, key: str) -> datetime:
        """Get required datetime value from row."""
        value = row.get(key)
        if value is None:
            raise WorkbookParseError(f"Missing required field: {key}")
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except (ValueError, TypeError):
            raise WorkbookParseError(f"Invalid datetime for field {key}: {value}")
    
    def _get_optional_datetime(self, row: Dict, key: str) -> Optional[datetime]:
        """Get optional datetime value from row."""
        value = row.get(key)
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except (ValueError, TypeError):
            return None
    
    def _parse_skill_level(self, row: Dict, column: str = "Skill Level") -> SkillLevel:
        """Parse skill level enum from row."""
        value = self._get_str(row, column)
        normalized = value.split("-")[-1].strip().lower() if "-" in value else value.strip().lower()
        mapping = {
            "junior": SkillLevel.JUNIOR,
            "beginner": SkillLevel.JUNIOR,
            "intermediate": SkillLevel.INTERMEDIATE,
            "mid": SkillLevel.MID,
            "senior": SkillLevel.SENIOR,
            "advanced": SkillLevel.ADVANCED,
            "expert": SkillLevel.EXPERT,
        }
        if normalized not in mapping:
            raise WorkbookParseError(f"Invalid skill level: {value}")
        return mapping[normalized]

    def _average_pct_columns(self, row: Dict, suffix: str) -> float:
        """Average percentage values for columns with a given suffix."""
        values = []
        for col_name, value in row.items():
            if col_name.endswith(suffix) and value is not None:
                try:
                    values.append(float(value))
                except (ValueError, TypeError):
                    continue
        return sum(values) / len(values) if values else 0.0
    
    def _parse_work_item_type(self, row: Dict) -> WorkItemType:
        """Parse work item type enum from row."""
        value = self._get_str(row, "Type")
        mapping = {
            "feature": WorkItemType.FEATURE,
            "story": WorkItemType.STORY,
            "task": WorkItemType.TASK,
            "bug": WorkItemType.BUG,
            "spike": WorkItemType.SPIKE,
            "defect": WorkItemType.DEFECT,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid work item type: {value}")
        return mapping[key]
    
    def _parse_priority(self, row: Dict) -> Priority:
        """Parse priority enum from row."""
        value = self._get_str(row, "Priority")
        mapping = {
            "critical": Priority.CRITICAL,
            "high": Priority.HIGH,
            "medium": Priority.MEDIUM,
            "low": Priority.LOW,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid priority: {value}")
        return mapping[key]
    
    def _parse_work_item_status(self, row: Dict) -> WorkItemStatus:
        """Parse work item status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "not started": WorkItemStatus.NOT_STARTED,
            "in progress": WorkItemStatus.IN_PROGRESS,
            "done": WorkItemStatus.DONE,
            "completed": WorkItemStatus.COMPLETED,
            "blocked": WorkItemStatus.BLOCKED,
            "spillover": WorkItemStatus.SPILLOVER,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid work item status: {value}")
        return mapping[key]
    
    def _parse_sprint_status(self, row: Dict) -> SprintStatus:
        """Parse sprint status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "not started": SprintStatus.NOT_STARTED,
            "in progress": SprintStatus.IN_PROGRESS,
            "completed": SprintStatus.COMPLETED,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid sprint status: {value}")
        return mapping[key]
    
    def _parse_blocker_severity(self, row: Dict) -> BlockerSeverity:
        """Parse blocker severity enum from row."""
        value = self._get_str(row, "Severity")
        mapping = {
            "critical": BlockerSeverity.CRITICAL,
            "high": BlockerSeverity.HIGH,
            "medium": BlockerSeverity.MEDIUM,
            "low": BlockerSeverity.LOW,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid blocker severity: {value}")
        return mapping[key]
    
    def _parse_blocker_status(self, row: Dict) -> BlockerStatus:
        """Parse blocker status enum from row."""
        value = self._get_str(row, "Status")
        mapping = {
            "open": BlockerStatus.OPEN,
            "resolved": BlockerStatus.RESOLVED,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid blocker status: {value}")
        return mapping[key]
    
    def _parse_dependency_type(self, row: Dict) -> DependencyType:
        """Parse dependency type enum from row."""
        value = self._get_str(row, "Dependency Type")
        mapping = {
            "finish-to-start": DependencyType.FINISH_TO_START,
            "start-to-start": DependencyType.START_TO_START,
        }
        key = value.lower()
        if key not in mapping:
            raise WorkbookParseError(f"Invalid dependency type: {value}")
        return mapping[key]
    
    def _parse_yes_no(self, value: Optional[str]) -> bool:
        """Parse Yes/No string to boolean."""
        if not value:
            return False
        return value.lower() in ["yes", "true", "1", "y"]
    
    def _normalize_sprint_name(self, sprint_name: Optional[str]) -> Optional[str]:
        """Normalize sprint name (e.g., 'Sprint 1' -> 'Sprint 1')."""
        if not sprint_name:
            return None
        return sprint_name.strip()
    
    def _generate_resource_id(self, resource_name: str) -> str:
        """Generate resource ID from name."""
        # Simple approach: use name as-is, could be improved
        return resource_name.lower().replace(" ", "_")
    
    def _generate_sprint_id(self, sprint_name: str, sprint_num: int) -> str:
        """Generate sprint ID from name and number."""
        # Format: SPR-1, SPR-2, etc.
        return f"SPR-{sprint_num}"
